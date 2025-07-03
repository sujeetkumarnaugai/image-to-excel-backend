from paddleocr import PaddleOCR
import openpyxl
import os

ocr = PaddleOCR(use_angle_cls=True, lang='en')

def convert_image_to_excel(image_path):
    result = ocr.ocr(image_path, cls=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Extracted text lines grouped as rows
    current_row = 1
    for line_group in result:
        for line in line_group:
            _, (text, _) = line
            columns = text.strip().split()
            for col_index, value in enumerate(columns, start=1):
                ws.cell(row=current_row, column=col_index, value=value)
            current_row += 1

    # Bold header row and auto-adjust width
    bold_font = openpyxl.styles.Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output_path = os.path.join("uploads", "converted.xlsx")
    wb.save(output_path)
    return output_path